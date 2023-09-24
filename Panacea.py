#!/usr/bin/env python
# coding: utf-8

# In[9]:


# import libraries
import pandas as pd
from pandas.api.types import is_object_dtype, is_numeric_dtype, is_bool_dtype, is_string_dtype, is_float_dtype
import numpy as np
import re
import os
import os.path
import traceback
from openpyxl import Workbook
import datetime
import argparse
from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.workbook.protection import WorkbookProtection


# In[6]:


# Import excel file and make updates on fOutputs sheets
def import_data(file_location):  
        
        wbook = load_workbook(file_location)
# Get the worksheets where we need to made corrections
        # Sheets where we have changes: RR6, RR7, RR8, fOut_RR, BIO2, SUP6, Validation, CWW9
        wsheet_RR = wbook['fOut_RR']  # Errata 7,8 and 9?
        wsheet_RR6 = wbook['RR6']  # Errata 13
        wsheet_RR7 = wbook['RR7']  # Errata 7
        wsheet_RR8 = wbook['RR8']  # Errata 8 and 9
        wsheet_RR20 = wbook['RR20']  # Errata 18
        wsheet_BIO2 = wbook['BIO2']  # Errata 4
        wsheet_SUP6 = wbook['SUP6']  # Errata 14
        wsheet_Validation = wbook['Validation']  # Errata 11 and 12
        wsheet_CWW3 = wbook['CWW3']  # Errata 15  TODO: This errata does not seem to have been implemented below
        wsheet_CWW9 = wbook['CWW9']  # Errata 3
        wsheet_CWW19 = wbook['CWW19']  # Errata 16
        wsheet_SUM2 = wbook['SUM2']  # Errata 20 and 21


# Assign specific values to cells
        wsheet_RR['J440'] = '=IF(ISBLANK(\'RR7\'!$E$47),"##BLANK",\'RR7\'!$E$47)'
        wsheet_RR['J443'] = '=IF(ISBLANK(\'RR7\'!$E$50),"##BLANK",\'RR7\'!$E$50)'
        wsheet_RR['J466'] = '=IF(ISBLANK(\'RR7\'!$E$81),"##BLANK",\'RR7\'!$E$81)'
        wsheet_RR['J512'] = '=IF(ISBLANK(\'RR8\'!$E$71),"##BLANK",\'RR8\'!$E$71)' 
        wsheet_RR7['W47'] = 'RR7_013_PR24'
        wsheet_RR7['W50'] = 'RR7_016_PR24'
        wsheet_RR7['W81'] = 'RR7_029_PR24'
        wsheet_RR8['W71'] = 'RR8_032_PR24'
        wsheet_BIO2['W71'] = 'RR8_032_PR24'   # Are we sure this correct?
        wsheet_Validation['C105'] = 'Wastewater network+ - WINEP nutrient removal (phosphorus and total nitrogen) scheme costs and cost drivers'
        wsheet_Validation['C107'] = 'Transition and accelerated programme - Wastewater network+ - Sewage treatment works population, capacity and network data'
        wsheet_SUP6['B15'] = 'New selective meter installation for existing customers'
        wsheet_SUP6['B16'] = 'New business meter installation for existing customers'
        wsheet_SUP6['X15'] = 'New selective meter installation for existing customers'
        wsheet_SUP6['X16'] = 'New business meter installation for existing customers'
        wsheet_SUM2['G45'] = '=SUM(G39:G44)'
        
# BIO2
        bypass_cells_BIO2=['L36', 'T36', 'AB36', 'AJ36', 'AR36', 'AZ36', 'BH36']
        for row in wsheet_BIO2['E36':'BO36']:
            for cell in row:
                if cell.coordinate in bypass_cells_BIO2:
                    continue
                else:
                    column_name = get_column_letter(cell.column)
                    # =IFERROR(E34+E35, 0)
                    cell.value = f"=IFERROR({column_name}34+{column_name}35,0)"
                
# remove #REF! from the formula
        for row in wsheet_CWW9['BW139':'CE139']:
            for cell in row:
                cell.value = re.sub('#REF!,', '', cell.value) 
                
# Update formula
        for row in wsheet_CWW19['E812':'K812']:
            for cell in row:
                column_name = get_column_letter(cell.column)
                # =IFERROR(SUM(E412:E811),0)
                cell.value = f"=IFERROR(SUM({column_name}412:{column_name}811),0)"
                
# Change cell background color
        # Create a PatternFill object with the desired color
        yellow_fill = PatternFill(start_color='FFEFCA', end_color='FFEFCA', fill_type='solid')
        grey_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        # Apply the background color to the cell
        wsheet_RR7['W47'].fill = yellow_fill
        wsheet_RR7['W50'].fill = yellow_fill
        wsheet_RR7['W81'].fill = yellow_fill
        wsheet_RR8['W71'].fill = yellow_fill
        wsheet_RR8['O71'].fill = grey_fill
        
        
        for row in wsheet_RR6['E22':'I27']: #set the range here 
            for cell in row:
                cell.fill = grey_fill  # Change cell colors
                
        for row in wsheet_RR6['Q22':'U27']:  # Set the range here
            for cell in row:
                cell.fill = grey_fill  # Change cell colors

        # Remove/Empty cells that previously had values
        for row in wsheet_RR['K443':'T443']:  # Set the range here
            for cell in row:
                cell.value = None  # Set a value or null here

        for row in wsheet_RR['K466':'T466']:
            for cell in row:
                cell.value = None 

        for row in wsheet_RR['K512':'T512']: 
            for cell in row:
                cell.value = None 
        # Remove values and change colour in a range of cells
        for row in wsheet_RR7['X50':'AG50']: 
            for cell in row:
                cell.value = None 
                cell.fill = grey_fill

        for row in wsheet_RR7['X81':'AG81']: 
            for cell in row:
                cell.value = None 
                cell.fill = grey_fill

        for row in wsheet_RR8['X71':'AG71']: 
            for cell in row:
                cell.value = None 
                cell.fill = grey_fill
                
        for row in wsheet_RR20['E10':'L10']: 
            for cell in row:
                cell.value = None 
                cell.fill = yellow_fill

        wsheet_RR['T512'].value = None

# Update Boncode
        for row in wsheet_RR20['V10':'AC10']: 
            for cell in row:
                cell.value = 'BO5002_PR24' 
                cell.fill = yellow_fill

        return wbook.save('PR24 BP tables V6-Publish.xlsx')


# In[10]:


# Compare two excel files for checking structural changes (if number of columns and rows are the same between two files)
def check_structure(file_location, V7_file_location):
    # Sheets where we have changes: RR6, RR7, RR8, fOut_RR, BIO2, SUP6, Validation, CWW9
    # Create Workbook object
    wb = Workbook()
    error_log_name = ("Auto_Fix_Error log_"+datetime.datetime.now().strftime("%d.%m.%Y %H.%M")+".xlsx")
    # Get active sheet
    ws = wb.active
    try: 
        # list_of_sheets=['CWW9', 'RR6', 'RR7', 'RR8', 'fOut_RR', 'BIO2', 'SUP6', 'Validation']
        wbook = load_workbook(file_location)
        wbook_V7 = load_workbook(V7_file_location)
        list_of_sheets = wbook_V7.sheetnames  # Get the list of worksheets in the file

        for sheet in list_of_sheets:
            row_error, column_error = (False, False)
            if wbook[sheet].max_row != wbook_V7[sheet].max_row:
                row_error = True
                message = (f"Max row number in {wbook[sheet].title} sheet is {wbook[sheet].max_row} and does not match the max row number in V7 template {wbook_V7[sheet].title} sheet which is {wbook_V7[sheet].max_row}")
                ws.append([message])        
            if wbook[sheet].max_column != wbook_V7[sheet].max_column:
                column_error= True
                message = (f"Max column number in {wbook[sheet].title} sheet is {wbook[sheet].max_column} and does not match the max column number in V7 template {wbook_V7[sheet].title} sheet which is {wbook_V7[sheet].max_column}")
                ws.append([message]) 
            if not row_error and not column_error:
                message = (f"Success, No Errors detected in {wbook[sheet].title} sheet")
                ws.append([message])    
    except Exception:
        print(traceback.format_exc())
        pass
    wb.save(error_log_name)


# In[11]:


get_ipython().run_cell_magic('time', '', "#####Main\n##Files Location\nfile_location ='PR24 BP tables V6-Publish.xlsx'\nV7_file_location ='Comparison_PR24.xlsx'\n##Functions\ncheck_structure(file_location,V7_file_location)\nwbook=import_data(file_location)")


# In[ ]:




